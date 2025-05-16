import os
import sys
import re
from typing import Dict, List, Optional, Union, Any
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext
import threading
import queue
import string

# Document processing imports
import PyPDF2
import csv
import json
import openpyxl
import requests
from bs4 import BeautifulSoup

# NLP imports
import nltk
from nltk.corpus import stopwords
from nltk.tokenize import word_tokenize
from nltk.tag import pos_tag
from nltk.stem import WordNetLemmatizer

# Search engine imports
import whoosh
from whoosh.index import create_in, open_dir
from whoosh.fields import Schema, TEXT, ID, STORED, NUMERIC
from whoosh.qparser import QueryParser, MultifieldParser
from whoosh.qparser import OrGroup, AndGroup
from whoosh.qparser.plugins import FuzzyTermPlugin, WildcardPlugin, PhrasePlugin
from whoosh.scoring import BM25F

# Download required NLTK data
try:
    nltk.data.find('tokenizers/punkt')
    nltk.data.find('corpora/stopwords')
    nltk.data.find('corpora/wordnet')
    nltk.data.find('taggers/averaged_perceptron_tagger')
except LookupError:
    nltk.download('punkt')
    nltk.download('stopwords')
    nltk.download('wordnet')
    nltk.download('averaged_perceptron_tagger')

def initialize_nltk():
    """Initialize NLTK by downloading required data."""
    required_nltk_data = [
        'punkt',
        'stopwords',
        'wordnet',
        'averaged_perceptron_tagger'
    ]
    
    print("Checking NLTK data...")
    for item in required_nltk_data:
        try:
            if item == 'wordnet':
                # Special case for wordnet
                try:
                    nltk.data.find(f'corpora/{item}')
                except LookupError:
                    print(f"Downloading {item}...")
                    nltk.download(item, quiet=True)
            else:
                try:
                    if item == 'punkt':
                        nltk.data.find(f'tokenizers/{item}')
                    else:
                        nltk.data.find(f'corpora/{item}')
                except LookupError:
                    print(f"Downloading {item}...")
                    nltk.download(item, quiet=True)
        except Exception as e:
            print(f"Error downloading {item}: {str(e)}")
            print("You may need to manually download NLTK data.")
            return False
    return True

class DocumentProcessor:
    def __init__(self, index_dir: str = "index"):
        """Initialize the document processor."""
        self.index_dir = index_dir
        
        # Initialize NLTK resources
        if not initialize_nltk():
            print("Warning: NLTK initialization incomplete. Some features may not work.")
            
        try:
            self.stop_words = set(stopwords.words('english'))
        except Exception as e:
            print(f"Warning: Could not load stopwords. Using empty set. Error: {str(e)}")
            self.stop_words = set()
            
        try:
            self.lemmatizer = WordNetLemmatizer()
        except Exception as e:
            print(f"Warning: Could not initialize lemmatizer. Using basic processing. Error: {str(e)}")
            self.lemmatizer = None
            
        self.create_index_dir()
        
        self.schema = Schema(
            path=ID(stored=True),
            filetype=ID(stored=True),
            location=ID(stored=True),
            content=TEXT(stored=True),
            score=NUMERIC(stored=True, sortable=True)
        )
        self.create_or_open_index()
        
    def create_index_dir(self):
        """Create index directory if it doesn't exist."""
        if not os.path.exists(self.index_dir):
            os.mkdir(self.index_dir)
            
    def create_or_open_index(self):
        """Create or open the search index."""
        if not os.listdir(self.index_dir):
            self.ix = create_in(self.index_dir, self.schema)
        else:
            self.ix = open_dir(self.index_dir)
            
    def preprocess_text(self, text: str) -> str:
        """Preprocess text according to requirements."""
        if not text:
            return ""

        try:
            # Lowercasing
            text = text.lower()

            # Tokenization
            try:
                tokens = word_tokenize(text)
            except Exception as e:
                print(f"Warning: Tokenization failed. Using basic split. Error: {str(e)}")
                tokens = text.split()

            # Remove punctuation and stopwords
            tokens = [token for token in tokens 
                     if token not in string.punctuation 
                     and (token not in self.stop_words if self.stop_words else True)]

            # POS tagging and Lemmatization
            if self.lemmatizer:
                try:
                    pos_tags = pos_tag(tokens)
                    
                    def get_wordnet_pos(tag):
                        if tag.startswith('J'):
                            return nltk.corpus.wordnet.ADJ
                        elif tag.startswith('V'):
                            return nltk.corpus.wordnet.VERB
                        elif tag.startswith('N'):
                            return nltk.corpus.wordnet.NOUN
                        elif tag.startswith('R'):
                            return nltk.corpus.wordnet.ADV
                        else:
                            return nltk.corpus.wordnet.NOUN

                    lemmatized = [self.lemmatizer.lemmatize(word, get_wordnet_pos(tag)) 
                                for word, tag in pos_tags]
                    return " ".join(lemmatized)
                except Exception as e:
                    print(f"Warning: Advanced preprocessing failed. Using basic tokens. Error: {str(e)}")
                    return " ".join(tokens)
            else:
                return " ".join(tokens)
        except Exception as e:
            print(f"Warning: Text preprocessing failed. Using raw text. Error: {str(e)}")
            return text.lower()

    def index_document(self, path: str, filetype: str, location: str, content: str):
        """Index a document or document part."""
        # Preprocess the content before indexing
        processed_content = self.preprocess_text(content)
        
        writer = self.ix.writer()
        writer.add_document(
            path=path,
            filetype=filetype,
            location=location,
            content=processed_content,
            score=1.0  # Default score, will be updated during search
        )
        writer.commit()
        
    def index_pdf(self, file_path: str) -> tuple[bool, str]:
        """Index a PDF file, page by page."""
        try:
            pdf_reader = PyPDF2.PdfReader(file_path)
            indexed_pages = 0
            
            for page_num, page in enumerate(pdf_reader.pages, 1):
                text = page.extract_text()
                if text.strip():
                    location = f"Page {page_num}"
                    self.index_document(file_path, "pdf", location, text)
                    indexed_pages += 1
                    
            if indexed_pages > 0:
                return True, f"Indexed PDF: {file_path} ({indexed_pages} pages)"
            return False, "PDF was empty"
        except Exception as e:
            return False, f"Error indexing PDF {file_path}: {str(e)}"
            
    def index_txt(self, file_path: str) -> tuple[bool, str]:
        """Index a text file."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                text = file.read()
                
            if text.strip():
                self.index_document(file_path, "txt", "Full document", text)
                return True, f"Indexed TXT: {file_path}"
            return False, "Text file was empty"
        except Exception as e:
            return False, f"Error indexing TXT {file_path}: {str(e)}"
            
    def index_csv(self, file_path: str) -> tuple[bool, str]:
        """Index a CSV file, row by row."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                csv_reader = csv.reader(file)
                headers = next(csv_reader, None)
                indexed_rows = 0
                
                for row_num, row in enumerate(csv_reader, 1):
                    if any(cell.strip() for cell in row):
                        content = " ".join(cell for cell in row if cell)
                        location = f"Row {row_num}"
                        self.index_document(file_path, "csv", location, content)
                        indexed_rows += 1
                        
            if indexed_rows > 0:
                return True, f"Indexed CSV: {file_path} ({indexed_rows} rows)"
            return False, "CSV was empty"
        except Exception as e:
            return False, f"Error indexing CSV {file_path}: {str(e)}"
            
    def index_excel(self, file_path: str) -> tuple[bool, str]:
        """Index an Excel file, row by row for each sheet."""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            total_indexed = 0
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                headers = [str(cell.value) if cell.value is not None else "" for cell in sheet[1]]
                
                for row_num, row in enumerate(sheet.iter_rows(min_row=2), 2):
                    row_values = [str(cell.value) if cell.value is not None else "" for cell in row]
                    if any(val.strip() for val in row_values):
                        content = " ".join(val for val in row_values if val)
                        location = f"Sheet '{sheet_name}', Row {row_num}"
                        self.index_document(file_path, "excel", location, content)
                        total_indexed += 1
                        
            if total_indexed > 0:
                return True, f"Indexed Excel: {file_path} ({total_indexed} rows across sheets)"
            return False, "Excel file was empty"
        except Exception as e:
            return False, f"Error indexing Excel {file_path}: {str(e)}"
            
    def index_json(self, file_path: str) -> tuple[bool, str]:
        """Flatten and index JSON structure."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
                data = json.load(file)
                
            def flatten_json(obj: Union[Dict, List], path: str = "") -> List[tuple[str, str]]:
                items = []
                
                if isinstance(obj, dict):
                    for key, value in obj.items():
                        new_path = f"{path}.{key}" if path else key
                        if isinstance(value, (dict, list)):
                            items.extend(flatten_json(value, new_path))
                        else:
                            items.append((new_path, str(value)))
                elif isinstance(obj, list):
                    for i, value in enumerate(obj):
                        new_path = f"{path}[{i}]"
                        if isinstance(value, (dict, list)):
                            items.extend(flatten_json(value, new_path))
                        else:
                            items.append((new_path, str(value)))
                            
                return items
                
            flattened = flatten_json(data)
            indexed_items = 0
            
            for path, value in flattened:
                if value.strip():
                    self.index_document(file_path, "json", path, value)
                    indexed_items += 1
                    
            if indexed_items > 0:
                return True, f"Indexed JSON: {file_path} ({indexed_items} items)"
            return False, "JSON was empty or contained no indexable content"
        except Exception as e:
            return False, f"Error indexing JSON {file_path}: {str(e)}"
            
    def index_web(self, url: str) -> tuple[bool, str]:
        """Index content from a web page."""
        try:
            response = requests.get(url, timeout=10)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            title = soup.title.text if soup.title else ""
            paragraphs = soup.find_all('p')
            
            # Index title and paragraphs separately
            indexed_items = 0
            
            if title.strip():
                self.index_document(url, "web", "Title", title)
                indexed_items += 1
                
            for i, p in enumerate(paragraphs, 1):
                text = p.get_text().strip()
                if text:
                    self.index_document(url, "web", f"Paragraph {i}", text)
                    indexed_items += 1
                    
            if indexed_items > 0:
                return True, f"Indexed Web Page: {url} ({indexed_items} sections)"
            return False, "Web page had no content"
        except Exception as e:
            return False, f"Error indexing web page {url}: {str(e)}"
    
    def search(self, query_text: str, file_type_filter: Optional[str] = None) -> List[Dict]:
        """Search the index with the given query."""
        # Preprocess the query the same way as documents
        processed_query = self.preprocess_text(query_text)
        
        with self.ix.searcher(weighting=BM25F()) as searcher:
            # Configure query parser with all required plugins
            qp = MultifieldParser(["content"], schema=self.ix.schema, group=OrGroup)
            
            # Add required query capabilities
            qp.add_plugin(FuzzyTermPlugin())  # For fuzzy queries
            qp.add_plugin(WildcardPlugin())   # For wildcard queries
            qp.add_plugin(PhrasePlugin())     # For phrase queries
            
            # Parse the query
            content_query = qp.parse(processed_query)
            
            # Create a combined query if there's a file type filter
            if file_type_filter:
                from whoosh import query
                type_query = query.Term("filetype", file_type_filter)
                final_query = query.And([content_query, type_query])
            else:
                final_query = content_query
                
            # Search with sorting by score
            results = searcher.search(final_query, limit=10, sortedby="score")
            
            return [{
                "path": hit["path"],
                "filetype": hit["filetype"],
                "location": hit["location"],
                "snippet": hit.highlights("content") or hit["content"][:150] + "...",
                "score": hit.score
            } for hit in results]
            
    def clear_index(self):
        """Clear the entire index."""
        if os.path.exists(self.index_dir):
            for file in os.listdir(self.index_dir):
                os.remove(os.path.join(self.index_dir, file))
            self.create_or_open_index()

class SearchEngineApp:
    def __init__(self, root):
        """Initialize the Search Engine GUI application."""
        self.root = root
        self.root.title("Multi-format Search Engine")
        self.root.geometry("900x800")


        self.processor = DocumentProcessor()
        self.setup_ui()
        self.task_queue = queue.Queue()
        self.is_working = False
        self.current_results = []
        
        # Start checking for completed tasks
        self.check_queue()
        
    def setup_ui(self):
        """Set up the user interface."""
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        search_frame = ttk.LabelFrame(main_frame, text="Search", padding=10)
        search_frame.pack(fill=tk.X, pady=5)
        
        # Query input
        query_frame = ttk.Frame(search_frame)
        query_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(query_frame, text="Search Query:").pack(side=tk.LEFT, padx=5)
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(query_frame, textvariable=self.search_var, width=50)
        search_entry.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        search_entry.bind("<Return>", lambda e: self.perform_search())
        
        ttk.Button(query_frame, text="Search", command=self.perform_search).pack(side=tk.LEFT, padx=5)
        
        # Filter frame
        filter_frame = ttk.Frame(search_frame)
        filter_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(filter_frame, text="File Types:").pack(side=tk.LEFT, padx=5)
        
        self.file_filters = {}
        for ft in ["PDF", "TXT", "CSV", "Excel", "JSON", "Web"]:
            var = tk.BooleanVar(value=True)
            self.file_filters[ft.lower()] = var
            ttk.Checkbutton(filter_frame, text=ft, variable=var).pack(side=tk.LEFT, padx=5)
        
        # Index section
        index_frame = ttk.LabelFrame(main_frame, text="Add Documents", padding=10)
        index_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(index_frame, text="Add Files", command=self.select_files).pack(side=tk.LEFT, padx=5)
        ttk.Button(index_frame, text="Add Folder", command=self.select_folder).pack(side=tk.LEFT, padx=5)
        
        # Web URL section
        ttk.Label(index_frame, text="Web URL:").pack(side=tk.LEFT, padx=5)
        self.url_var = tk.StringVar()
        ttk.Entry(index_frame, textvariable=self.url_var, width=30).pack(side=tk.LEFT, padx=5)
        ttk.Button(index_frame, text="Add URL", command=self.index_url).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(index_frame, text="Clear All", command=self.clear_index).pack(side=tk.RIGHT, padx=5)
        
        # Results section
        results_frame = ttk.LabelFrame(main_frame, text="Search Results")
        results_frame.pack(fill=tk.X, expand=False, pady=5)

        self.results_text = scrolledtext.ScrolledText(results_frame, wrap=tk.WORD, height=8)
        self.results_text.pack(fill=tk.X, expand=False)

        # Configure tags for result formatting
        self.results_text.tag_configure("title", font=("Arial", 10, "bold"))
        self.results_text.tag_configure("location", font=("Arial", 9, "italic"))
        self.results_text.tag_configure("score", font=("Arial", 9))
        self.results_text.tag_configure("snippet", font=("Arial", 10))
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN).pack(fill=tk.X, pady=5)


        # Evaluation section
        eval_frame = ttk.LabelFrame(main_frame, text="Evaluation", padding=10)
        eval_frame.pack(fill=tk.X, pady=5)
        
        # Relevant results input
        input_frame = ttk.Frame(eval_frame)
        input_frame.pack(fill=tk.X, pady=5)
        
        ttk.Label(input_frame, text="Total Relevant Documents:").pack(side=tk.LEFT, padx=5)
        self.total_relevant_var = tk.StringVar(value="0")
        ttk.Entry(input_frame, textvariable=self.total_relevant_var, width=10).pack(side=tk.LEFT, padx=5)
        
        ttk.Label(input_frame, text="Relevant in Results:").pack(side=tk.LEFT, padx=5)
        self.relevant_found_var = tk.StringVar(value="0")
        ttk.Entry(input_frame, textvariable=self.relevant_found_var, width=10).pack(side=tk.LEFT, padx=5)
        
        ttk.Button(input_frame, text="Calculate Metrics", command=self.calculate_metrics).pack(side=tk.LEFT, padx=5)
        
        # Metrics display
        metrics_frame = ttk.Frame(eval_frame)
        metrics_frame.pack(fill=tk.X, pady=5)
        
        # Precision
        precision_frame = ttk.Frame(metrics_frame)
        precision_frame.pack(side=tk.LEFT, padx=20)
        ttk.Label(precision_frame, text="Precision:").pack(side=tk.LEFT)
        self.precision_var = tk.StringVar(value="N/A")
        ttk.Label(precision_frame, textvariable=self.precision_var).pack(side=tk.LEFT, padx=5)
        
        # Recall
        recall_frame = ttk.Frame(metrics_frame)
        recall_frame.pack(side=tk.LEFT, padx=20)
        ttk.Label(recall_frame, text="Recall:").pack(side=tk.LEFT)
        self.recall_var = tk.StringVar(value="N/A")
        ttk.Label(recall_frame, textvariable=self.recall_var).pack(side=tk.LEFT, padx=5)
        
        # F1-score
        f1_frame = ttk.Frame(metrics_frame)
        f1_frame.pack(side=tk.LEFT, padx=20)
        ttk.Label(f1_frame, text="F1-score:").pack(side=tk.LEFT)
        self.f1_var = tk.StringVar(value="N/A")
        ttk.Label(f1_frame, textvariable=self.f1_var).pack(side=tk.LEFT, padx=5)

        
    def update_status(self, message: str):
        """Update the status bar message."""
        self.status_var.set(message)
        self.root.update_idletasks()


    def calculate_metrics(self):
        """Calculate precision, recall, and F1-score."""
        try:
            # Get values from input fields
            total_relevant = int(self.total_relevant_var.get())
            relevant_found = int(self.relevant_found_var.get())
            retrieved = len(self.current_results)
            
            if retrieved == 0:
                self.update_status("No search results to evaluate.")
                return
                
            if total_relevant < relevant_found:
                self.update_status("Error: Relevant found cannot be greater than total relevant documents.")
                return
                
            # Calculate metrics
            precision = relevant_found / retrieved if retrieved > 0 else 0
            recall = relevant_found / total_relevant if total_relevant > 0 else 0
            f1 = 2 * (precision * recall) / (precision + recall) if (precision + recall) > 0 else 0
            
            # Update display
            self.precision_var.set(f"{precision:.3f}")
            self.recall_var.set(f"{recall:.3f}")
            self.f1_var.set(f"{f1:.3f}")
            
            self.update_status("Metrics calculated successfully.")
            
        except ValueError:
            self.update_status("Please enter valid numbers for relevant documents.")
        except Exception as e:
            self.update_status(f"Error calculating metrics: {str(e)}")

        
    def update_results(self, results: List[Dict]):
        """Update the results area with search results."""
        self.results_text.delete(1.0, tk.END)
        
        if not results:
            self.results_text.insert(tk.END, "No results found.")
            return
            
        for i, result in enumerate(results, 1):
            # File name and type
            path = os.path.basename(result["path"]) if result["filetype"] != "web" else result["path"]
            self.results_text.insert(tk.END, f"{i}. {path} ({result['filetype']})\n", "title")
            
            # Location
            self.results_text.insert(tk.END, f"   Location: {result['location']}\n", "location")
            
            # Score
            self.results_text.insert(tk.END, f"   Score: {result['score']:.4f}\n", "score")
            
            # Snippet
            self.results_text.insert(tk.END, f"   {result['snippet']}\n\n", "snippet")
        
    def perform_search(self):
        """Execute search query and display results."""
        query = self.search_var.get().strip()
        if not query:
            self.update_status("Please enter a search query.")
            return
            
        # Get selected file types
        selected_types = [ft for ft, var in self.file_filters.items() if var.get()]
        if not selected_types:
            self.update_status("Please select at least one file type.")
            return
            
        self.update_status(f"Searching for: {query}")
        
        # Search in each selected file type
        all_results = []
        for ft in selected_types:
            results = self.processor.search(query, ft)
            all_results.extend(results)
            
        # Sort by score
        all_results.sort(key=lambda x: x["score"], reverse=True)
        self.current_results = all_results
        self.precision_var.set("N/A")
        self.recall_var.set("N/A")
        self.f1_var.set("N/A")
        self.total_relevant_var.set("0")
        self.relevant_found_var.set("0")
        
        # Update display
        self.update_results(all_results)
        self.update_status(f"Found {len(all_results)} results.")
        
    def select_files(self):
        """Select and index multiple files."""
        file_paths = filedialog.askopenfilenames(
            title="Select files to index",
            filetypes=[
                ("All supported", "*.pdf;*.txt;*.csv;*.xlsx;*.json"),
                ("PDF files", "*.pdf"),
                ("Text files", "*.txt"),
                ("CSV files", "*.csv"),
                ("Excel files", "*.xlsx"),
                ("JSON files", "*.json"),
                ("All files", "*.*")
            ]
        )
        
        if file_paths:
            self.start_worker_thread(self.index_files, file_paths)
        
    def select_folder(self):
        """Select and index all files in a folder."""
        folder_path = filedialog.askdirectory(title="Select folder to index")
        
        if folder_path:
            file_paths = []
            for root, _, files in os.walk(folder_path):
                for file in files:
                    if file.lower().endswith((".pdf", ".txt", ".csv", ".xlsx", ".json")):
                        file_paths.append(os.path.join(root, file))
                        
            if file_paths:
                self.start_worker_thread(self.index_files, file_paths)
            else:
                self.update_status("No supported files found in the folder.")
        
    def index_url(self):
        """Index a web URL."""
        url = self.url_var.get().strip()
        if not url:
            self.update_status("Please enter a URL.")
            return
            
        if not url.startswith(("http://", "https://")):
            url = "http://" + url
            
        self.start_worker_thread(self.processor.index_web, url)
        
    def index_files(self, file_paths):
        """Index multiple files."""
        total_files = len(file_paths)
        successful = 0
        
        for i, file_path in enumerate(file_paths, start=1):
            self.update_status(f"Indexing file {i}/{total_files}: {os.path.basename(file_path)}")
            
            try:
                extension = os.path.splitext(file_path)[1].lower()
                
                if extension == ".pdf":
                    success, message = self.processor.index_pdf(file_path)
                elif extension == ".txt":
                    success, message = self.processor.index_txt(file_path)
                elif extension == ".csv":
                    success, message = self.processor.index_csv(file_path)
                elif extension == ".xlsx":
                    success, message = self.processor.index_excel(file_path)
                elif extension == ".json":
                    success, message = self.processor.index_json(file_path)
                else:
                    success, message = False, f"Unsupported file type: {extension}"
                    
                if success:
                    successful += 1
                    
                self.task_queue.put(message)
                
            except Exception as e:
                self.task_queue.put(f"Error processing {file_path}: {str(e)}")
                
        return f"Indexed {successful}/{total_files} files successfully."
        
    def clear_index(self):
        """Clear the search index."""
        self.processor.clear_index()
        self.update_status("Index cleared.")
        
    def start_worker_thread(self, target_func, *args):
        """Start a worker thread for long-running operations."""
        if self.is_working:
            self.update_status("A task is already in progress. Please wait.")
            return
            
        self.is_working = True
        self.worker_thread = threading.Thread(
            target=self.run_task,
            args=(target_func, *args)
        )
        self.worker_thread.daemon = True
        self.worker_thread.start()
        
    def run_task(self, target_func, *args):
        """Run a task in a separate thread and put the result in the queue."""
        try:
            result = target_func(*args)
            self.task_queue.put(result)
        except Exception as e:
            self.task_queue.put(f"Error: {str(e)}")
        finally:
            self.is_working = False
            
    def check_queue(self):
        """Check for completed tasks and update the UI."""
        try:
            while True:
                message = self.task_queue.get_nowait()
                self.update_status(message)
        except queue.Empty:
            pass
            
        self.root.after(100, self.check_queue)

def main():
    try:
        root = tk.Tk()
        app = SearchEngineApp(root)
        root.mainloop()
    except Exception as e:
        print(f"Error starting application: {str(e)}")
        input("Press Enter to exit...")

if __name__ == "__main__":
    main() 